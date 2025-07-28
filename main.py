import openpyxl
import os
import configparser
import json
import requests
from typing import Optional, List, Dict, Any
from datetime import datetime
import logging

# Configurazione del logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', filename='app.log', filemode='w')
logger = logging.getLogger(__name__)

# Variabili globali
DEFAULT_FILE_PATH: str = "data.xlsx"
API_CONFIG: Dict[str, str] = {}
AUTH_TOKEN: Optional[str] = None
MODIFY_EXCEL: bool = False

def main() -> None:
    workbook = open_excel_file(DEFAULT_FILE_PATH)
    if workbook is None:
        logger.error("Impossibile aprire il file Excel")
        return
    
    sheet = workbook.active
    
    # Leggere i dati e creare il JSON
    data = read_excel_data(sheet)
    json_output = create_json_body(data)
    
    # Stampare il JSON
    logger.info("JSON creato:")
    logger.info(json.dumps(json_output, indent=2, ensure_ascii=False))

    # Salvare il JSON su file (opzionale)
    save_json_to_file(json_output, "output.json")
    
    # Eseguire il workflow API se la configurazione è disponibile
    if API_CONFIG and API_CONFIG['base_url']:
        execute_api_workflow(json_output)
    else:
        logger.warning("Configurazione API non disponibile. Saltando le chiamate HTTP.")

    workbook.close()

def read_excel_data(sheet) -> List[Dict[str, Any]]:
    """Legge i dati dal foglio Excel e li converte in una lista di dizionari"""
    data = []
    
    # Assumendo che la prima riga contenga le intestazioni
    headers = []
    visible_columns = []
    
    for col in range(1, 20):  # Aumentato il range per controllare più colonne
        # Verificare se la colonna è nascosta
        column_letter = sheet.cell(row=1, column=col).column_letter
        column_dimension = sheet.column_dimensions.get(column_letter)
        
        # Se la colonna è nascosta, saltarla
        if column_dimension and column_dimension.hidden:
            continue
            
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(str(cell_value))
            visible_columns.append(col)
        elif not cell_value and col <= 6:  # Solo per le prime 6 colonne
            break

    logger.info(f"Intestazioni trovate: {headers}")
    logger.info(f"Colonne visibili: {visible_columns}")

    # Leggere i dati dalle righe successive
    row = 2
    while True:
        # Controllare se la riga è vuota
        if sheet.cell(row=row, column=3).value is None:
            break
            
        record = {}
        for i, header in enumerate(headers):
            col = visible_columns[i]  # Usare l'indice corretto della colonna visibile
            cell_value = sheet.cell(row=row, column=col).value
            
            if header == "Importo":
                # Convertire l'importo in valore positivo
                if cell_value is not None:
                    try:
                        # Rimuovere il segno negativo se presente
                        importo = abs(float(cell_value))
                        record[header] = importo
                    except (ValueError, TypeError):
                        record[header] = 0.0
                else:
                    record[header] = 0.0
            elif header == "Data":
                # Gestire la data
                if cell_value is not None:
                    if isinstance(cell_value, datetime):
                        record[header] = cell_value.strftime("%d/%m/%Y")
                    else:
                        record[header] = str(cell_value)
                else:
                    record[header] = ""
            else:
                # Altri campi come stringa
                record[header] = str(cell_value) if cell_value is not None else ""
        
        data.append(record)
        row += 1
    
    logger.info(f"Trovati {len(data)} record")
    
    if MODIFY_EXCEL:
        # Modifica l'Excel se richiesto
        data = modify_excel(data)
        logger.info(f"Dati dopo la modifica: {len(data)} record")
        
    return data

def create_json_body(data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Crea il corpo JSON per la chiamata API"""
    
    # Calcolare il totale degli importi
    total_amount = sum(record.get("Importo", 0) for record in data)
    
    json_body = {
        "transactions": data,
        "summary": {
            "total_records": len(data),
            "total_amount": round(total_amount, 2),
            "processed_date": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        }
    }
    
    return json_body

def save_json_to_file(data: Dict[str, Any], filename: str) -> None:
    """Salva il JSON su file"""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"JSON salvato in: {filename}")
    except Exception as e:
        logger.error(f"Errore nel salvare il JSON: {e}")

def login_api() -> bool:
    """Effettua il login e ottiene il token di autenticazione"""
    global AUTH_TOKEN
    
    if not API_CONFIG:
        logger.warning("Configurazione API non caricata")
        return False
    
    login_url = API_CONFIG['base_url'] + API_CONFIG['login_endpoint']
    login_data = {
        "email": API_CONFIG['email'],
        "password": API_CONFIG['password']
    }
    
    try:
        logger.info(f"Tentativo di login su: {login_url}")
        response = requests.post(
            login_url,
            json=login_data,
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        
        if response.status_code == 200:
            response_data = response.json()
            # Assumendo che il token sia nel campo 'token' o 'access_token'
            AUTH_TOKEN = response_data["data"].get('token') or response_data["data"].get('access_token')

            if AUTH_TOKEN:
                logger.info("Login effettuato con successo")
                return True
            else:
                logger.warning("Token non trovato nella risposta del login")
                return False
        else:
            logger.error(f"Errore nel login: {response.status_code} - {response.text}")
            return False
            
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore di connessione durante il login: {e}")
        return False

def api_get_request(endpoint_override: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Effettua una chiamata GET autenticata"""
    global AUTH_TOKEN
    
    if not AUTH_TOKEN:
        logger.warning("Token di autenticazione non disponibile. Effettuare prima il login.")
        return None
    
    endpoint = endpoint_override or API_CONFIG['get_endpoint']
    url = API_CONFIG['base_url'] + endpoint
    
    headers = {
        'Authorization': f'Bearer {AUTH_TOKEN}',
        'Content-Type': 'application/json'
    }
    
    try:
        logger.info(f"Chiamata GET a: {url}")
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            logger.info("Chiamata GET completata con successo")
            return response.json()
        elif response.status_code == 401:
            logger.warning("Token scaduto o non valido. Rieffettuare il login.")
            AUTH_TOKEN = None
            return None
        else:
            logger.error(f"Errore nella chiamata GET: {response.status_code} - {response.text}")
            return None
            
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore di connessione nella chiamata GET: {e}")
        return None

def api_post_request(data: Dict[str, Any], endpoint_override: Optional[str] = None) -> bool:
    """Effettua una chiamata POST autenticata"""
    global AUTH_TOKEN
    
    if not AUTH_TOKEN:
        logger.warning("Token di autenticazione non disponibile. Effettuare prima il login.")
        return False
    
    endpoint = endpoint_override or API_CONFIG['post_endpoint']
    url = API_CONFIG['base_url'] + endpoint
    
    headers = {
        'Authorization': f'Bearer {AUTH_TOKEN}',
        'Content-Type': 'application/json'
    }
    
    body = {
        "name": data["Titolo"],
        "description": data["Descrizione"],
        "amount": data["Importo"],
        "type": data["Tipo"],
        "date": data["Data"]
    }
    
    try:
        logger.info(f"Chiamata POST a: {url}")
        response = requests.post(url, json=body, headers=headers, timeout=30)
        
        if response.status_code in [200, 201]:
            logger.info("Chiamata POST completata con successo")
            return True
        elif response.status_code == 401:
            logger.warning("Token scaduto o non valido. Rieffettuare il login.")
            AUTH_TOKEN = None
            return False
        else:
            logger.error(f"Errore nella chiamata POST: {response.status_code} - {response.text}")
            return False
            
    except requests.exceptions.RequestException as e:
        logger.error(f"Errore di connessione nella chiamata POST: {e}")
        return False

def execute_api_workflow(json_data: Dict[str, Any]) -> None:
    """Esegue il workflow completo: login, GET, POST"""
    logger.info("=== Inizio workflow API ===")
    
    # 1. Login
    if not login_api():
        logger.error("Impossibile effettuare il login. Workflow interrotto.")
        return
    
    # 2. Chiamata GET (opzionale - per verificare dati esistenti)
    logger.info("\n--- Chiamata GET ---")
    types_response = api_get_request()
    type_data = types_response.get("data") if types_response else None
    if type_data:
        logger.info(f"Dati esistenti trovati: {len(type_data)} record")
        
    # Prima di effettuare la POST, modifica il tipo spesa da stringa ad _id
    for item in json_data.get("transactions", []):
        tipo_spesa = item.get("Tipo")
        if tipo_spesa and type_data:
            tipo_record = next((t for t in type_data if t.get("name") == tipo_spesa), None)
            if tipo_record:
                item["Tipo"] = tipo_record.get("_id")

    # 3. Chiamata POST con i dati del JSON
    logger.info("\n--- Chiamata POST ---")
    try: 
        transactional_data = json_data.get("transactions", [])
        for item in transactional_data:
            logger.debug(item)
            if api_post_request(item):
                logger.info("Dati inviati con successo!")
            else:
                logger.error("Errore nell'invio dei dati")
    except Exception as e:
        logger.error(f"Errore durante la preparazione dei dati per la POST: {e}")

    logger.info("=== Fine workflow API ===\n")

# open an excel file
def open_excel_file(file_path: str) -> Optional[openpyxl.Workbook]:
    try:
        # data_only=True fa sì che vengano letti i valori calcolati invece delle formule
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        logger.info("Excel file opened successfully.")
        return workbook
    except Exception as e:
        logger.error(f"Error opening Excel file: {e}")
        return None
    
def setup() -> None:
    global DEFAULT_FILE_PATH, API_CONFIG, MODIFY_EXCEL
    config_file_path = "file_da_caricare.ini"
    
    if os.path.exists(config_file_path):
        config = configparser.ConfigParser()
        config.read(config_file_path)
        
        # Leggere il percorso del file
        if 'EXCEL' in config and 'file_path' in config['EXCEL']:
            DEFAULT_FILE_PATH = config['EXCEL']['file_path']
            MODIFY_EXCEL = config['EXCEL'].getboolean('modify', False)
            logger.info(f"Using file path from config: {DEFAULT_FILE_PATH}")
        
        # Leggere la configurazione API
        if 'API' in config:
            API_CONFIG = {
                'base_url': config['API'].get('base_url', ''),
                'login_endpoint': config['API'].get('login_endpoint', ''),
                'get_endpoint': config['API'].get('get_endpoint', ''),
                'post_endpoint': config['API'].get('post_endpoint', ''),
                'email': config['API'].get('email', ''),
                'password': config['API'].get('password', '')
            }
            logger.info(f"API Config loaded: {API_CONFIG['base_url']}")
        else:
            logger.warning("Sezione [API] non trovata nel file di configurazione")
    else:
        logger.error("File di configurazione non trovato")

def modify_excel(data_input: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Funzione per modificare l'Excel, se necessario"""
    # Questa funzione può essere implementata per modificare il file Excel
    # in base alle esigenze specifiche del workflow.
    
    # rimuovi gli oggetti che non sono necessari, come quelli che nel titolo o nella descrizione contengono "american express"
    filtered_data = [
        item for item in data_input
        if "american express" not in item.get("Titolo", "").lower()
        and "american express" not in item.get("Descrizione", "").lower()
    ]
    
    return filtered_data

if __name__ == "__main__":
    setup()
    main()